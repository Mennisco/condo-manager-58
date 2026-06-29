"""One-off: set per-unit late fees (25% of dues, from Fee Incr. Calc. H27-H36)
and import 2023-2025 monthly fee history from the user's spreadsheet.
Idempotent: skips fee rows that already exist for (unit_id, year, month).
Does NOT touch 2026 data."""
import os
from datetime import timezone
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from pymongo import MongoClient

load_dotenv(Path(__file__).parent / ".env")
client = MongoClient(os.environ["MONGO_URL"])
db = client[os.environ["DB_NAME"]]

XLSX = "/tmp/innsbruck.xlsx"
wb = openpyxl.load_workbook(XLSX, data_only=True)

# Map address number -> unit doc
units = list(db.units.find())
by_addr = {str(int(float(u["unit_number"]))): u for u in units}

# 1) Late fees from Fee Incr. Calc. current-year block (rows 27-36, B=addr, H=late fee)
calc = wb["Fee Incr. Calc."]
lf_set = 0
for r in range(27, 37):
    addr = calc.cell(row=r, column=2).value  # B
    lf = calc.cell(row=r, column=8).value    # H
    if addr is None or lf is None:
        continue
    key = str(int(float(addr)))
    u = by_addr.get(key)
    if not u:
        continue
    db.units.update_one({"_id": u["_id"]}, {"$set": {"late_fee": round(float(lf), 2)}})
    lf_set += 1
print(f"late fees set on {lf_set} units")

# 2) Import history. Blocks: (year, first_data_row, last_data_row)
blocks = [(2023, 3, 12), (2024, 16, 25), (2025, 32, 41)]
log = wb["Mo. Fees Log"]
created = skipped = 0
for year, r0, r1 in blocks:
    for r in range(r0, r1 + 1):
        addr = log.cell(row=r, column=2).value  # B
        amt_due = log.cell(row=r, column=4).value  # D
        if addr is None:
            continue
        key = str(int(float(addr)))
        u = by_addr.get(key)
        if not u:
            continue
        amt = float(amt_due) if amt_due is not None else float(u.get("monthly_fee", 0))
        for m in range(1, 13):
            col = 6 + (m - 1)  # F=6 .. Q=17
            cell = log.cell(row=r, column=col).value
            exists = db.fee_payments.find_one(
                {"unit_id": str(u["_id"]), "period_year": year, "period_month": m}
            )
            if exists:
                skipped += 1
                continue
            if cell is not None and hasattr(cell, "year"):
                paid_date = cell.replace(tzinfo=timezone.utc).isoformat()
                doc = {
                    "unit_id": str(u["_id"]),
                    "unit_number": u["unit_number"],
                    "owner_name": u["owner_name"],
                    "period_year": year,
                    "period_month": m,
                    "amount_due": amt,
                    "amount_paid": amt,
                    "paid": True,
                    "paid_date": paid_date,
                    "method": None,
                    "late_fee_waived": False,
                    "notes": None,
                    "created_at": paid_date,
                }
            else:
                doc = {
                    "unit_id": str(u["_id"]),
                    "unit_number": u["unit_number"],
                    "owner_name": u["owner_name"],
                    "period_year": year,
                    "period_month": m,
                    "amount_due": amt,
                    "amount_paid": 0.0,
                    "paid": False,
                    "paid_date": None,
                    "method": None,
                    "late_fee_waived": False,
                    "notes": None,
                    "created_at": None,
                }
            db.fee_payments.insert_one(doc)
            created += 1
print(f"history: created={created} skipped(existing)={skipped}")
print("2023:", db.fee_payments.count_documents({"period_year": 2023}))
print("2024:", db.fee_payments.count_documents({"period_year": 2024}))
print("2025:", db.fee_payments.count_documents({"period_year": 2025}))
print("2026:", db.fee_payments.count_documents({"period_year": 2026}))
