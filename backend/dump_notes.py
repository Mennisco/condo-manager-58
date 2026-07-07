import openpyxl
wb = openpyxl.load_workbook('/tmp/innsbruck.xlsx')  # comments need default (not data_only)
ws = wb['Mo. Fees Log']
print("dims:", ws.dimensions, "max_row", ws.max_row, "max_col", ws.max_column)
# Print all cells that have comments
found = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.comment is not None:
            found += 1
            txt = cell.comment.text.replace("\n", " | ")
            print(f"{cell.coordinate} (r{cell.row},c{cell.column}) val={cell.value!r} :: NOTE: {txt}")
print("total comments:", found)
