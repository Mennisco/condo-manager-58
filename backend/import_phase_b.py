"""Phase B import: ownership % onto units, Paid Expenses (2023-2025) and
Budgets (2024-2027) from the user's spreadsheet. Idempotent via a source tag."""
import os
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from pymongo import MongoClient

load_dotenv(Path(__file__).parent / ".env")
client = MongoClient(os.environ["MONGO_URL"])
db = client[os.environ["DB_NAME"]]
wb = openpyxl.load_workbook("/tmp/innsbruck.xlsx", data_only=True)

units = list(db.units.find())
by_addr = {str(int(float(u["unit_number"]))): u for u in units}

# 1) ownership % from Fee Incr. Calc. rows 27-36 (B=addr, D=ownership)
calc = wb["Fee Incr. Calc."]
n = 0
for r in range(27, 37):
    addr = calc.cell(row=r, column=2).value
    pct = calc.cell(row=r, column=4).value
    if addr is None or pct is None:
        continue
    u = by_addr.get(str(int(float(addr))))
    if u:
        db.units.update_one({"_id": u["_id"]}, {"$set": {"ownership_pct": round(float(pct), 4)}})
        n += 1
print(f"ownership % set on {n} units")

# 2) Paid Expenses -> normalize categories
CAT = {
    "mow": "Mowing", "lawn mowing": "Mowing",
    "snow": "Snow Removal", "snow removal": "Snow Removal",
    "util": "Utilities", "utilities": "Utilities",
    "ins": "Insurance", "insurance": "Insurance",
    "acctg": "Bank/Accounting", "bank/accounting": "Bank/Accounting",
    "maint": "Maintenance", "maintenance": "Maintenance",
    "ww": "Window Washing", "window washing": "Window Washing",
    "landscaping": "Landscaping", "trash removal": "Trash Removal",
    "other": "Other",
}
exp_blocks = [(2023, 3, 10), (2024, 18, 26), (2025, 33, 41), (2026, 48, 56), (2027, 63, 71)]
pe = wb["Paid Expenses"]
wb_c = openpyxl.load_workbook("/tmp/innsbruck.xlsx")  # with cell comments
pe_c = wb_c["Paid Expenses"]
db.expenses.delete_many({"source": "xlsx_paid_expenses"})
created = 0
for year, r0, r1 in exp_blocks:
    for r in range(r0, r1 + 1):
        label = pe.cell(row=r, column=1).value
        if not label or str(label).strip().upper().startswith("TOTAL"):
            continue
        category = CAT.get(str(label).strip().lower(), str(label).strip().title())
        for m in range(1, 13):
            col = 1 + m  # B=2 (Jan) .. M=13 (Dec)
            val = pe.cell(row=r, column=col).value
            comment = pe_c.cell(row=r, column=col).comment
            note_text = comment.text.strip() if comment and comment.text else None
            if val is None:
                # A note may exist on a cell with no numeric amount — skip (no expense to attach to)
                continue
            try:
                amt = float(val)
            except (TypeError, ValueError):
                continue
            if amt == 0:
                continue
            db.expenses.insert_one({
                "date": f"{year}-{m:02d}-01",
                "category": category,
                "vendor": None,
                "description": f"{category} — {year}-{m:02d}",
                "amount": round(amt, 2),
                "method": None,
                "date_paid": None,
                "notes": note_text or "Imported from spreadsheet",
                "source": "xlsx_paid_expenses",
                "created_at": None,
            })
            created += 1
print(f"expenses imported: {created}")

# 3) Budgets -> one budget_item per category per year (N col total)
bud_blocks = [(2024, 5, 14), (2025, 22, 30), (2026, 39, 48), (2027, 55, 63)]
bd = wb["Budget"]
bcreated = bupdated = 0
for year, r0, r1 in bud_blocks:
    for r in range(r0, r1 + 1):
        cat = bd.cell(row=r, column=1).value
        total = bd.cell(row=r, column=14).value  # N
        if not cat:
            continue
        try:
            amt = round(float(total), 2) if total is not None else 0.0
        except (TypeError, ValueError):
            amt = 0.0
        category = str(cat).strip()
        existing = db.budget_items.find_one({"year": year, "category": category})
        if existing:
            db.budget_items.update_one({"_id": existing["_id"]}, {"$set": {"budgeted_amount": amt}})
            bupdated += 1
        else:
            db.budget_items.insert_one({
                "year": year,
                "category": category,
                "budgeted_amount": amt,
                "notes": "Imported from spreadsheet",
                "created_at": None,
            })
            bcreated += 1
print(f"budgets: created={bcreated} updated={bupdated}")
for y in (2023, 2024, 2025, 2026):
    print(f"expenses {y}:", db.expenses.count_documents({"date": {"$gte": f"{y}-01-01", "$lte": f"{y}-12-31"}}))
for y in (2024, 2025, 2026, 2027):
    print(f"budget {y}:", db.budget_items.count_documents({"year": y}))
